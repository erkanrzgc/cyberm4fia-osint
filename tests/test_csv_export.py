"""CSV + XLSX export tests."""

from __future__ import annotations

import csv
import io
import zipfile
from pathlib import Path

import pytest

from core.models import EmailResult, PlatformResult, ScanResult
from core.reporter import export_csv, export_xlsx, xlsx_available
from modules.crypto.models import CryptoIntel
from modules.phone.models import PhoneIntel


def _result() -> ScanResult:
    r = ScanResult(username="alice")
    r.scan_time = 1.25
    r.platforms = [
        PlatformResult(
            platform="GitHub",
            url="https://github.com/alice",
            category="dev",
            exists=True,
            status="found",
            response_time=0.123,
            http_status=200,
            confidence=0.9,
        ),
        PlatformResult(
            platform="Twitter",
            url="https://twitter.com/alice",
            category="social",
            exists=False,
            status="not_found",
        ),
    ]
    r.emails = [
        EmailResult(
            email="alice@example.com",
            source="gravatar",
            verified=True,
            breach_count=2,
            breaches=["LinkedIn2012", "Adobe2013"],
        )
    ]
    r.phone_intel = [
        PhoneIntel(
            raw="+14155552671",
            e164="+14155552671",
            country_code=1,
            region="US",
            country_name="United States",
            carrier="AT&T",
            line_type="mobile",
            timezones=("America/Los_Angeles",),
            valid=True,
            sources=("phonenumbers",),
        )
    ]
    r.crypto_intel = [
        CryptoIntel(
            address="0x742d35Cc6634C0532925a3b844Bc9e7595f0bEb0",
            chain="eth",
            balance=1.5,
            tx_count=10,
            source="etherscan",
        )
    ]
    return r


# ── CSV ─────────────────────────────────────────────────────────────


def _read_csv(zf: zipfile.ZipFile, name: str) -> list[list[str]]:
    with zf.open(name) as fp:
        text = io.TextIOWrapper(fp, encoding="utf-8", newline="")
        return list(csv.reader(text))


def test_export_csv_writes_zip_bundle(tmp_path: Path) -> None:
    path = tmp_path / "report.csv"
    export_csv(_result(), str(path))

    zip_path = path.with_suffix(".zip")
    assert zip_path.exists()

    with zipfile.ZipFile(zip_path) as zf:
        names = set(zf.namelist())
        assert {"summary.csv", "platforms.csv", "emails.csv",
                "phones.csv", "crypto.csv", "geo.csv"} <= names

        summary = _read_csv(zf, "summary.csv")
        assert summary[0] == ["field", "value"]
        assert ["username", "alice"] in summary
        assert ["found_count", "1"] in summary

        platforms = _read_csv(zf, "platforms.csv")
        assert platforms[0][:4] == ["platform", "category", "url", "exists"]
        github_row = next(row for row in platforms if row[0] == "GitHub")
        assert github_row[3] == "true"  # exists
        assert github_row[4] == "found"

        emails = _read_csv(zf, "emails.csv")
        assert emails[1][0] == "alice@example.com"
        assert "LinkedIn2012" in emails[1][5]


def test_export_csv_accepts_explicit_zip_suffix(tmp_path: Path) -> None:
    path = tmp_path / "bundle.csv.zip"
    export_csv(_result(), str(path))
    assert path.exists()
    with zipfile.ZipFile(path) as zf:
        assert "platforms.csv" in zf.namelist()


def test_export_csv_handles_empty_sections(tmp_path: Path) -> None:
    path = tmp_path / "blank.csv"
    export_csv(ScanResult(username="ghost"), str(path))
    with zipfile.ZipFile(path.with_suffix(".zip")) as zf:
        platforms = _read_csv(zf, "platforms.csv")
        # header only
        assert len(platforms) == 1


# ── XLSX ────────────────────────────────────────────────────────────


@pytest.mark.skipif(not xlsx_available(), reason="openpyxl not installed")
def test_export_xlsx_writes_one_sheet_per_section(tmp_path: Path) -> None:
    from openpyxl import load_workbook

    path = tmp_path / "report.xlsx"
    export_xlsx(_result(), str(path))
    assert path.exists()

    wb = load_workbook(path)
    assert {"summary", "platforms", "emails", "phones",
            "crypto", "geo"} <= set(wb.sheetnames)

    platforms = wb["platforms"]
    header = [c.value for c in platforms[1]]
    assert header[:3] == ["platform", "category", "url"]
    # Two data rows (GitHub + Twitter).
    assert platforms.max_row == 3
